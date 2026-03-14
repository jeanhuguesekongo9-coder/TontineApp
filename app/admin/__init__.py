# -*- coding: utf-8 -*-
from flask import Blueprint, render_template, redirect, url_for, flash, request, Response
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime
from ..models import db, Utilisateur, KYC, Tontine, Paiement, AuditLog, Notification, MembreTontine
import openpyxl
import io

admin = Blueprint("admin", __name__)

def admin_requis(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.role != "admin":
            flash("Acces reserve aux administrateurs.", "danger")
            return redirect(url_for("main.tableau_de_bord"))
        return f(*args, **kwargs)
    return decorated

@admin.route("/")
@admin_requis
def dashboard():
    stats = {
        "total_utilisateurs": Utilisateur.query.count(),
        "kyc_en_attente": KYC.query.filter_by(statut="en_attente").count(),
        "kyc_approuves": KYC.query.filter_by(statut="approuve").count(),
        "tontines_actives": Tontine.query.filter_by(statut="en_cours").count(),
        "tontines_recrut": Tontine.query.filter_by(statut="recrutement").count(),
        "paiements_en_attente": Paiement.query.filter_by(statut="en_attente").count(),
    }
    kycs_recents = KYC.query.filter_by(statut="en_attente").order_by(KYC.created_at.desc()).limit(10).all()
    return render_template("admin/dashboard.html", stats=stats, kycs_recents=kycs_recents)

@admin.route("/kyc")
@admin_requis
def gerer_kyc():
    statut_filtre = request.args.get("statut", "en_attente")
    kycs = KYC.query.filter_by(statut=statut_filtre).order_by(KYC.created_at).all()
    return render_template("admin/kyc.html", kycs=kycs, statut_filtre=statut_filtre)

@admin.route("/kyc/<int:kyc_id>/approuver", methods=["POST"])
@admin_requis
def approuver_kyc(kyc_id):
    kyc = KYC.query.get_or_404(kyc_id)
    kyc.statut = "approuve"
    kyc.verifie_par = current_user.id
    kyc.verifie_le = datetime.utcnow()
    kyc.utilisateur.kyc_valide = True
    notif = Notification(user_id=kyc.user_id,
        titre="Dossier approuve",
        message="Votre dossier KYC a ete approuve. Vous pouvez rejoindre une tontine !",
        type_notif="success",
        lien=url_for("tontines.liste"))
    db.session.add(notif)
    db.session.commit()
    AuditLog.log(current_user.id, "kyc_approuve", f"User: {kyc.user_id}", ip=request.remote_addr)
    db.session.commit()
    flash("KYC approuve.", "success")
    return redirect(url_for("admin.gerer_kyc"))

@admin.route("/kyc/<int:kyc_id>/rejeter", methods=["POST"])
@admin_requis
def rejeter_kyc(kyc_id):
    kyc = KYC.query.get_or_404(kyc_id)
    note = request.form.get("note", "").strip()
    kyc.statut = "rejete"
    kyc.note_admin = note
    kyc.verifie_par = current_user.id
    notif = Notification(user_id=kyc.user_id,
        titre="Dossier rejete",
        message=f"Votre dossier KYC a ete rejete. Motif : {note}.",
        type_notif="danger",
        lien=url_for("kyc.soumettre"))
    db.session.add(notif)
    db.session.commit()
    flash("KYC rejete.", "warning")
    return redirect(url_for("admin.gerer_kyc"))

@admin.route("/tontines")
@admin_requis
def gerer_tontines():
    tontines = Tontine.query.order_by(Tontine.created_at.desc()).all()
    return render_template("admin/tontines.html", tontines=tontines)

@admin.route("/tontines/creer", methods=["GET", "POST"])
@admin_requis
def creer_tontine():
    paniers = [50000, 100000, 200000]
    if request.method == "POST":
        nom = request.form.get("nom", "").strip()
        montant = float(request.form.get("montant_panier", 0))
        min_m = int(request.form.get("min_membres", 5))
        max_m = int(request.form.get("max_membres", 10))
        jour = int(request.form.get("jour_collecte", 1))
        description = request.form.get("description", "")
        if not nom or montant not in paniers:
            flash("Donnees invalides.", "danger")
            return render_template("admin/creer_tontine.html", paniers=paniers)
        tontine = Tontine(code=Tontine.generer_code(), nom=nom,
            montant_panier=montant, min_membres=min_m, max_membres=max_m,
            jour_collecte=jour, description=description, createur_id=current_user.id)
        db.session.add(tontine)
        db.session.commit()
        flash(f"Tontine {tontine.nom} creee !", "success")
        return redirect(url_for("admin.gerer_tontines"))
    return render_template("admin/creer_tontine.html", paniers=paniers)

@admin.route("/tontines/<int:tontine_id>/demarrer", methods=["POST"])
@admin_requis
def demarrer_tontine(tontine_id):
    import random
    tontine = Tontine.query.get_or_404(tontine_id)
    if not tontine.peut_demarrer:
        flash("Pas assez de membres.", "warning")
        return redirect(url_for("admin.gerer_tontines"))
    tontine.statut = "en_cours"
    tontine.date_debut = datetime.utcnow()
    membres = tontine.membres.all()
    random.shuffle(membres)
    for i, m in enumerate(membres, 1):
        m.ordre_collecte = i
    db.session.commit()
    flash(f"Tontine {tontine.nom} demarree !", "success")
    return redirect(url_for("admin.gerer_tontines"))

@admin.route("/utilisateurs")
@admin_requis
def utilisateurs():
    users = Utilisateur.query.order_by(Utilisateur.created_at.desc()).all()
    return render_template("admin/utilisateurs.html", users=users)

@admin.route("/logs")
@admin_requis
def logs():
    page = request.args.get("page", 1, type=int)
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).paginate(page=page, per_page=50)
    return render_template("admin/logs.html", logs=logs)

@admin.route("/export-excel")
@admin_requis
def export_excel():
    wb = openpyxl.Workbook()
    header_fill = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    def style_headers(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    # Onglet Utilisateurs
    ws1 = wb.active
    ws1.title = "Utilisateurs"
    style_headers(ws1, ["ID", "Email", "Role", "Email verifie", "KYC valide", "Compte actif", "Date inscription"])
    for u in Utilisateur.query.order_by(Utilisateur.created_at.desc()).all():
        ws1.append([u.id, u.email, u.role, u.email_verifie, u.kyc_valide, u.compte_actif, str(u.created_at)[:19]])

    # Onglet Profils
    ws2 = wb.create_sheet("Profils")
    style_headers(ws2, ["ID", "User ID", "Nom", "Prenom", "Telephone", "Ville", "Pays", "Profession"])
    from ..models import Profil
    for p in Profil.query.all():
        ws2.append([p.id, p.user_id, p.nom, p.prenom, p.telephone, p.ville or "", p.pays or "", p.profession or ""])

    # Onglet KYC
    ws3 = wb.create_sheet("KYC")
    style_headers(ws3, ["ID", "User ID", "Type doc", "Statut", "Note admin", "Date soumission"])
    for k in KYC.query.order_by(KYC.created_at.desc()).all():
        ws3.append([k.id, k.user_id, k.type_doc or "", k.statut, k.note_admin or "", str(k.created_at)[:19]])

    # Onglet Tontines
    ws4 = wb.create_sheet("Tontines")
    style_headers(ws4, ["ID", "Code", "Nom", "Montant panier", "Min membres", "Max membres", "Nb membres", "Statut", "Date creation"])
    for t in Tontine.query.order_by(Tontine.created_at.desc()).all():
        ws4.append([t.id, t.code, t.nom, t.montant_panier, t.min_membres, t.max_membres, t.nombre_membres, t.statut, str(t.created_at)[:19]])

    # Onglet Membres Tontines
    ws5 = wb.create_sheet("Membres Tontines")
    style_headers(ws5, ["ID", "User ID", "Tontine ID", "Ordre collecte", "A recu", "Statut", "Date adhesion"])
    for m in MembreTontine.query.all():
        ws5.append([m.id, m.user_id, m.tontine_id, m.ordre_collecte or "", m.a_recu, m.statut, str(m.date_adhesion)[:19]])

    # Onglet Paiements
    ws6 = wb.create_sheet("Paiements")
    style_headers(ws6, ["ID", "Reference", "User ID", "Tontine ID", "Montant", "Type", "Statut", "Date"])
    for p in Paiement.query.order_by(Paiement.created_at.desc()).all():
        ws6.append([p.id, p.reference, p.user_id, p.tontine_id, p.montant, p.type_paiement or "", p.statut, str(p.created_at)[:19]])

    # Ajuster largeurs
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"tontinesecure_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@admin.route("/fichier/<path:chemin>")
@admin_requis
def voir_fichier(chemin):
    import os
    from supabase import create_client
    from flask import redirect
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    client = create_client(url, key)
    result = client.storage.from_("kyc-documents").create_signed_url(chemin, 300)
    return redirect(result["signedURL"])
@admin.route("/recharges")
@admin_requis
def gerer_recharges():
    from ..models import Recharge
    statut = request.args.get("statut", "en_attente")
    recharges = Recharge.query.filter_by(statut=statut).order_by(Recharge.created_at.desc()).all()
    RESEAUX = {
        "wave_sn": "Wave Senegal", "wave_ci": "Wave CI",
        "orange_sn": "Orange Senegal", "orange_ci": "Orange CI", "mtn_ci": "MTN CI"
    }
    return render_template("admin/recharges.html", recharges=recharges, statut=statut, RESEAUX=RESEAUX)

@admin.route("/recharges/<int:recharge_id>/valider", methods=["POST"])
@admin_requis
def valider_recharge(recharge_id):
    from ..models import Recharge, Solde, Transaction
    recharge = Recharge.query.get_or_404(recharge_id)
    if recharge.statut != "en_attente":
        flash("Cette recharge a deja ete traitee.", "warning")
        return redirect(url_for("admin.gerer_recharges"))
    solde = Solde.query.filter_by(user_id=recharge.user_id).first()
    if not solde:
        solde = Solde(user_id=recharge.user_id, montant=0.0)
        db.session.add(solde)
        db.session.flush()
    solde_avant = solde.montant
    solde.montant += recharge.montant
    recharge.statut = "valide"
    recharge.valide_par = current_user.id
    recharge.valide_le = datetime.utcnow()
    trx = Transaction(
        reference=Transaction.generer_reference(),
        user_id=recharge.user_id,
        type_transaction="recharge",
        montant=recharge.montant,
        sens="credit",
        solde_avant=solde_avant,
        solde_apres=solde.montant,
        description=f"Recharge via {recharge.reseau.upper()} - Ref: {recharge.reference_transaction}"
    )
    db.session.add(trx)
    notif = Notification(
        user_id=recharge.user_id,
        titre="Recharge validee !",
        message=f"Votre recharge de {recharge.montant:,.0f} FCFA a ete validee. Nouveau solde: {solde.montant:,.0f} FCFA",
        type_notif="success",
        lien=url_for("paiements.dashboard")
    )
    db.session.add(notif)
    db.session.commit()
    AuditLog.log(current_user.id, "recharge_validee", f"User: {recharge.user_id}, Montant: {recharge.montant}", ip=request.remote_addr)
    db.session.commit()
    flash(f"Recharge de {recharge.montant:,.0f} FCFA validee !", "success")
    return redirect(url_for("admin.gerer_recharges"))

@admin.route("/recharges/<int:recharge_id>/rejeter", methods=["POST"])
@admin_requis
def rejeter_recharge(recharge_id):
    from ..models import Recharge
    recharge = Recharge.query.get_or_404(recharge_id)
    note = request.form.get("note", "").strip()
    recharge.statut = "rejete"
    recharge.note_admin = note
    recharge.valide_par = current_user.id
    recharge.valide_le = datetime.utcnow()
    notif = Notification(
        user_id=recharge.user_id,
        titre="Recharge rejetee",
        message=f"Votre recharge de {recharge.montant:,.0f} FCFA a ete rejetee. Motif: {note}",
        type_notif="danger",
        lien=url_for("paiements.recharger")
    )
    db.session.add(notif)
    db.session.commit()
    flash("Recharge rejetee.", "warning")
    return redirect(url_for("admin.gerer_recharges"))

@admin.route("/soldes")
@admin_requis
def gerer_soldes():
    from ..models import Solde
    soldes = Solde.query.order_by(Solde.montant.desc()).all()
    total = sum(s.montant for s in soldes)
    return render_template("admin/soldes.html", soldes=soldes, total=total)
@admin.route("/penalites")
@admin_requis
def gerer_penalites():
    from ..models import Penalite, Utilisateur, Tontine
    statut = request.args.get("statut", "en_cours")
    penalites = Penalite.query.filter_by(statut=statut).order_by(Penalite.created_at.desc()).all()
    total_du = sum(p.montant_du + p.montant_penalite for p in penalites if p.statut == "en_cours")
    mises_en_demeure = Penalite.query.filter_by(statut="mise_en_demeure").count()
    return render_template("admin/penalites.html",
        penalites=penalites, statut=statut,
        total_du=total_du, mises_en_demeure=mises_en_demeure)

@admin.route("/penalites/<int:pen_id>/regulariser", methods=["POST"])
@admin_requis
def regulariser_penalite(pen_id):
    from ..models import Penalite, Solde, Transaction, MembreTontine
    pen = Penalite.query.get_or_404(pen_id)
    solde = Solde.query.filter_by(user_id=pen.user_id).first()
    montant_total = pen.montant_du + pen.montant_penalite
    if not solde or solde.montant < montant_total:
        flash(f"Solde insuffisant. Necessite {montant_total:,.0f} FCFA.", "danger")
        return redirect(url_for("admin.gerer_penalites"))
    solde_avant = solde.montant
    solde.montant -= montant_total
    trx = Transaction(
        reference=Transaction.generer_reference(),
        user_id=pen.user_id,
        tontine_id=pen.tontine_id,
        type_transaction="cotisation",
        montant=montant_total,
        sens="debit",
        solde_avant=solde_avant,
        solde_apres=solde.montant,
        description=f"Regularisation penalite {pen.reference} - cotisation {pen.mois_reference} + penalite 1.5%"
    )
    db.session.add(trx)
    pen.statut = "regularise"
    pen.regularise_le = datetime.utcnow()
    membre = MembreTontine.query.filter_by(user_id=pen.user_id, tontine_id=pen.tontine_id).first()
    if membre and membre.statut == "suspendu":
        autres_pen = Penalite.query.filter_by(user_id=pen.user_id, tontine_id=pen.tontine_id, statut="mise_en_demeure").count()
        if autres_pen == 0:
            membre.statut = "actif"
    notif = Notification(
        user_id=pen.user_id,
        titre="Penalite regularisee",
        message=f"Votre penalite de {pen.mois_reference} a ete regularisee. Montant debite: {montant_total:,.0f} FCFA (cotisation + 1.5%). Solde restant: {solde.montant:,.0f} FCFA",
        type_notif="success",
        lien="/paiements/"
    )
    db.session.add(notif)
    AuditLog.log(current_user.id, "penalite_regularisee", f"Penalite {pen.reference}, user {pen.user_id}", ip=request.remote_addr)
    db.session.commit()
    flash(f"Penalite regularisee ! {montant_total:,.0f} FCFA debites.", "success")
    return redirect(url_for("admin.gerer_penalites"))

@admin.route("/penalites/<int:pen_id>/annuler", methods=["POST"])
@admin_requis
def annuler_penalite(pen_id):
    from ..models import Penalite, MembreTontine
    pen = Penalite.query.get_or_404(pen_id)
    note = request.form.get("note", "Annulee par admin").strip()
    pen.statut = "annulee"
    pen.note_admin = note
    pen.regularise_le = datetime.utcnow()
    membre = MembreTontine.query.filter_by(user_id=pen.user_id, tontine_id=pen.tontine_id).first()
    if membre and membre.statut == "suspendu":
        membre.statut = "actif"
    notif = Notification(
        user_id=pen.user_id,
        titre="Penalite annulee",
        message=f"Votre penalite du mois {pen.mois_reference} a ete annulee par l administration. Motif: {note}",
        type_notif="success",
        lien="/paiements/"
    )
    db.session.add(notif)
    AuditLog.log(current_user.id, "penalite_annulee", f"Penalite {pen.reference}, motif: {note}", ip=request.remote_addr)
    db.session.commit()
    flash("Penalite annulee.", "warning")
    return redirect(url_for("admin.gerer_penalites"))

@admin.route("/tableau-de-bord")
@admin_requis
def tableau_de_bord():
    from ..models import Solde, Recharge, Transaction, Penalite, MembreTontine, Tontine, Utilisateur, KYC
    total_soldes = db.session.query(db.func.sum(Solde.montant)).scalar() or 0
    total_membres = Utilisateur.query.filter_by(role="membre").count()
    membres_actifs = MembreTontine.query.filter_by(statut="actif").count()
    membres_suspendus = MembreTontine.query.filter_by(statut="suspendu").count()
    recharges_attente = Recharge.query.filter_by(statut="en_attente").count()
    from ..models import Retrait
    retraits_attente = Retrait.query.filter_by(statut="en_attente").count()
    montant_retraits_attente = db.session.query(db.func.sum(Retrait.montant)).filter_by(statut="en_attente").scalar() or 0
    penalites_cours = Penalite.query.filter_by(statut="en_cours").count()
    mises_en_demeure = Penalite.query.filter_by(statut="mise_en_demeure").count()
    total_penalites_dues = db.session.query(db.func.sum(Penalite.montant_penalite)).filter(
        Penalite.statut.in_(["en_cours", "mise_en_demeure"])
    ).scalar() or 0
    total_cotisations = db.session.query(db.func.sum(Transaction.montant)).filter_by(
        type_transaction="cotisation", sens="debit"
    ).scalar() or 0
    tontines_actives = Tontine.query.filter_by(statut="en_cours").count()
    kyc_attente = KYC.query.filter_by(statut="en_attente").count()
    retardataires = db.session.query(
        Penalite.user_id,
        Utilisateur.email,
        db.func.count(Penalite.id).label("nb_impayes"),
        db.func.sum(Penalite.montant_du + Penalite.montant_penalite).label("total_du")
    ).join(Utilisateur, Penalite.user_id == Utilisateur.id).filter(
        Penalite.statut.in_(["en_cours", "mise_en_demeure"])
    ).group_by(Penalite.user_id, Utilisateur.email).order_by(
        db.func.count(Penalite.id).desc()
    ).limit(10).all()
    from ..models import Profil
    retardataires_detail = []
    for r in retardataires:
        profil = Profil.query.filter_by(user_id=r.user_id).first()
        retardataires_detail.append({
            "user_id": r.user_id,
            "nom": profil.nom_complet if profil else r.email,
            "email": r.email,
            "nb_impayes": r.nb_impayes,
            "total_du": r.total_du,
            "en_poursuite": r.nb_impayes >= 3
        })
    return render_template("admin/tableau_de_bord.html",
        total_soldes=total_soldes,
        total_membres=total_membres,
        membres_actifs=membres_actifs,
        membres_suspendus=membres_suspendus,
        recharges_attente=recharges_attente,
        retraits_attente=retraits_attente,
        montant_retraits_attente=montant_retraits_attente,
        penalites_cours=penalites_cours,
        mises_en_demeure=mises_en_demeure,
        total_penalites_dues=total_penalites_dues,
        total_cotisations=total_cotisations,
        tontines_actives=tontines_actives,
        kyc_attente=kyc_attente,
        retardataires=retardataires_detail
    )
@admin.route("/contrats")
@admin_requis
def gerer_contrats():
    from ..models import Contrat, Utilisateur, Profil
    statut = request.args.get("statut", "signes")
    if statut == "signes":
        contrats = Contrat.query.filter_by(signe=True).order_by(Contrat.signe_le.desc()).all()
    else:
        contrats = Contrat.query.filter_by(signe=False).order_by(Contrat.created_at.desc()).all()
    total_signes = Contrat.query.filter_by(signe=True).count()
    total_en_attente = Contrat.query.filter_by(signe=False).count()
    return render_template("admin/contrats.html",
        contrats=contrats,
        statut=statut,
        total_signes=total_signes,
        total_en_attente=total_en_attente
    )

@admin.route("/contrats/voir/<string:reference>")
@admin_requis
def voir_contrat_admin(reference):
    from ..models import Contrat
    contrat = Contrat.query.filter_by(reference=reference).first_or_404()
    if not contrat.contenu_html:
        flash("Ce contrat n'a pas encore de contenu archivé.", "warning")
        return redirect(url_for("admin.gerer_contrats"))
    return contrat.contenu_html, 200, {"Content-Type": "text/html; charset=utf-8"}
@admin.route("/retraits")
@admin_requis
def gerer_retraits():
    from ..models import Retrait, Profil
    statut = request.args.get("statut", "en_attente")
    retraits = Retrait.query.filter_by(statut=statut).order_by(Retrait.created_at.desc()).all()
    total_en_attente = Retrait.query.filter_by(statut="en_attente").count()
    montant_en_attente = db.session.query(db.func.sum(Retrait.montant)).filter_by(statut="en_attente").scalar() or 0
    total_valides = Retrait.query.filter_by(statut="valide").count()
    montant_valide = db.session.query(db.func.sum(Retrait.montant)).filter_by(statut="valide").scalar() or 0
    reseaux = {
        "wave_sn": "Wave Sénégal",
        "wave_ci": "Wave Côte d'Ivoire",
        "orange_sn": "Orange Money Sénégal",
        "orange_ci": "Orange Money CI",
        "mtn_ci": "MTN CI"
    }
    return render_template("admin/retraits.html",
        retraits=retraits, statut=statut,
        total_en_attente=total_en_attente,
        montant_en_attente=montant_en_attente,
        total_valides=total_valides,
        montant_valide=montant_valide,
        reseaux=reseaux
    )

@admin.route("/retraits/<int:retrait_id>/valider", methods=["POST"])
@admin_requis
def valider_retrait(retrait_id):
    from ..models import Retrait, Solde, Transaction, Notification, AuditLog
    retrait = Retrait.query.get_or_404(retrait_id)
    if retrait.statut != "en_attente":
        flash("Ce retrait a déjà été traité.", "warning")
        return redirect(url_for("admin.gerer_retraits"))
    reference_envoi = request.form.get("reference_envoi", "").strip()
    if not reference_envoi:
        flash("La référence de transaction est obligatoire pour valider.", "danger")
        return redirect(url_for("admin.gerer_retraits"))
    # Débiter le solde du membre
    solde = Solde.query.filter_by(user_id=retrait.user_id).first()
    if not solde or solde.montant < retrait.montant:
        flash(f"Solde insuffisant pour ce membre ({solde.montant:,.0f} FCFA disponible).", "danger")
        return redirect(url_for("admin.gerer_retraits"))
    solde_avant = solde.montant
    solde.montant -= retrait.montant
    # Enregistrer la transaction
    trx = Transaction(
        reference=Transaction.generer_reference(),
        user_id=retrait.user_id,
        type_transaction="retrait",
        montant=retrait.montant,
        sens="debit",
        solde_avant=solde_avant,
        solde_apres=solde.montant,
        description=f"Retrait {retrait.reference} via {retrait.reseau} vers {retrait.numero_telephone} — Réf. envoi : {reference_envoi}"
    )
    db.session.add(trx)
    # Mettre à jour le retrait
    retrait.statut = "valide"
    retrait.reference_admin = reference_envoi
    retrait.traite_le = datetime.utcnow()
    # Notifier le membre
    reseaux_noms = {
        "wave_sn": "Wave Sénégal", "wave_ci": "Wave Côte d'Ivoire",
        "orange_sn": "Orange Money Sénégal", "orange_ci": "Orange Money CI", "mtn_ci": "MTN CI"
    }
    nom_reseau = reseaux_noms.get(retrait.reseau, retrait.reseau)
    notif = Notification(
        user_id=retrait.user_id,
        titre="Retrait envoyé ✅",
        message=f"Votre retrait de {retrait.montant:,.0f} FCFA a été envoyé via {nom_reseau} vers le {retrait.numero_telephone}. Référence : {reference_envoi}. Nouveau solde : {solde.montant:,.0f} FCFA.",
        type_notif="success",
        lien="/paiements/mes-retraits"
    )
    db.session.add(notif)
    AuditLog.log(current_user.id, "retrait_valide",
        f"Retrait {retrait.reference}, {retrait.montant:,.0f} FCFA, user {retrait.user_id}, réf. {reference_envoi}",
        ip=request.remote_addr)
    db.session.commit()
    flash(f"Retrait de {retrait.montant:,.0f} FCFA validé ! Solde membre débité.", "success")
    return redirect(url_for("admin.gerer_retraits"))

@admin.route("/retraits/<int:retrait_id>/rejeter", methods=["POST"])
@admin_requis
def rejeter_retrait(retrait_id):
    from ..models import Retrait, Notification, AuditLog
    retrait = Retrait.query.get_or_404(retrait_id)
    if retrait.statut != "en_attente":
        flash("Ce retrait a déjà été traité.", "warning")
        return redirect(url_for("admin.gerer_retraits"))
    motif = request.form.get("motif", "Rejeté par l'administrateur").strip()
    retrait.statut = "rejete"
    retrait.note_admin = motif
    retrait.traite_le = datetime.utcnow()
    notif = Notification(
        user_id=retrait.user_id,
        titre="Retrait rejeté",
        message=f"Votre demande de retrait de {retrait.montant:,.0f} FCFA a été rejetée. Motif : {motif}. Votre solde n'a pas été débité.",
        type_notif="danger",
        lien="/paiements/mes-retraits"
    )
    db.session.add(notif)
    AuditLog.log(current_user.id, "retrait_rejete",
        f"Retrait {retrait.reference}, motif : {motif}",
        ip=request.remote_addr)
    db.session.commit()
    flash("Retrait rejeté. Le membre en a été notifié.", "warning")
    return redirect(url_for("admin.gerer_retraits"))
